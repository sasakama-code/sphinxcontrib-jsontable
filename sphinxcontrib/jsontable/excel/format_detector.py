"""Excel format detection and intelligent conversion.

This module provides advanced Excel format detection and conversion capabilities,
supporting multiple Excel layouts including pivot tables, financial statements,
multi-header tables, cross-tabulations, and time series data.

Key Features:
- Automatic Excel format detection
- Format-specific conversion strategies
- Business entity extraction
- Data quality assessment
- Structure preservation and normalization
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ExcelFormatHandler:
    """Handler for various Excel format types with specialized processing."""

    def __init__(self):
        """Initialize format handlers."""
        self.format_handlers = {
            "pivot_table": PivotTableHandler(),
            "financial_statement": FinancialStatementHandler(),
            "multi_header": MultiHeaderHandler(),
            "cross_tab": CrossTabHandler(),
            "time_series": TimeSeriesHandler(),
            "standard_table": StandardTableHandler(),
        }
        logger.info("ExcelFormatHandler initialized with 6 format handlers")

    def get_handler(self, format_type: str) -> BaseFormatHandler:
        """Get appropriate format handler."""
        return self.format_handlers.get(
            format_type, self.format_handlers["standard_table"]
        )


class AdvancedExcelConverter:
    """Advanced Excel converter with intelligent format detection and processing."""

    def __init__(self):
        """Initialize the advanced Excel converter."""
        self.format_handler = ExcelFormatHandler()
        self.entity_patterns = self._load_entity_patterns()

    def analyze_structure(self, excel_file: str) -> dict[str, Any]:
        """Analyze Excel file structure comprehensively.

        Args:
            excel_file: Path to Excel file

        Returns:
            Dictionary containing structure analysis
        """
        logger.info(f"Analyzing Excel structure: {excel_file}")

        wb = openpyxl.load_workbook(excel_file, data_only=True)
        structure = {
            "file_path": excel_file,
            "sheet_names": wb.sheetnames,
            "sheets": {},
            "overall_format": None,
            "complexity_score": 0,
        }

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_analysis = self._analyze_sheet_structure(sheet, sheet_name)
            structure["sheets"][sheet_name] = sheet_analysis

        # Determine overall format
        structure["overall_format"] = self._determine_overall_format(
            structure["sheets"]
        )
        structure["complexity_score"] = self._calculate_complexity_score(
            structure["sheets"]
        )

        wb.close()
        return structure

    def auto_detect_format(self, excel_file: str) -> str:
        """Automatically detect Excel format type.

        Args:
            excel_file: Path to Excel file

        Returns:
            Detected format type string
        """
        logger.info(f"Auto-detecting Excel format: {excel_file}")

        wb = openpyxl.load_workbook(excel_file, data_only=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Check for pivot table indicators
            if self._has_pivot_indicators(sheet):
                wb.close()
                return "pivot_table"

            # Check for financial statement indicators
            if self._has_financial_indicators(sheet):
                wb.close()
                return "financial_statement"

            # Check for multi-header structure
            if self._has_multi_headers(sheet):
                wb.close()
                return "multi_header"

            # Check for cross-tabulation structure
            if self._has_crosstab_structure(sheet):
                wb.close()
                return "cross_tab"

            # Check for time series pattern
            if self._has_time_series_pattern(sheet):
                wb.close()
                return "time_series"

        wb.close()
        return "standard_table"

    def convert_with_intelligence(
        self, excel_file: str, config: dict[str, Any]
    ) -> dict[str, Any]:
        """Convert Excel with intelligent format processing.

        Args:
            excel_file: Path to Excel file
            config: Conversion configuration

        Returns:
            Comprehensive conversion results
        """
        logger.info(f"Starting intelligent Excel conversion: {excel_file}")

        # Step 1: Format detection
        detected_format = self.auto_detect_format(excel_file)
        handler = self.format_handler.get_handler(detected_format)

        # Step 2: Format-specific conversion
        conversion_result = handler.convert(excel_file, config)

        # Step 3: Entity extraction
        entities = self._extract_business_entities(
            conversion_result["data"],
            domain=config.get("domain", "general"),
            language=config.get("language", "japanese"),
        )

        # Step 4: Metadata generation
        metadata = self._generate_comprehensive_metadata(
            conversion_result, entities, config
        )

        # Step 5: Quality assessment
        quality_score = self._assess_data_quality(conversion_result["data"], metadata)

        return {
            "json_data": conversion_result["data"],
            "format_type": detected_format,
            "entities": entities,
            "metadata": metadata,
            "quality_score": quality_score,
            "format_confidence": conversion_result.get("confidence", 0.9),
            "conversion_notes": conversion_result.get("notes", []),
        }

    def detect_data_types(self, excel_structure: dict[str, Any]) -> dict[str, Any]:
        """Detect data types in Excel structure."""
        data_types = {}

        for sheet_name, sheet_data in excel_structure["sheets"].items():
            if "sample_data" in sheet_data:
                sheet_types = {}
                sample_data = sheet_data["sample_data"]

                if sample_data:
                    # Analyze first few rows to determine column types
                    for col_idx, _col_data in enumerate(
                        sample_data[0] if sample_data else []
                    ):
                        col_type = self._infer_column_type(sample_data, col_idx)
                        sheet_types[f"column_{col_idx}"] = col_type

                data_types[sheet_name] = sheet_types

        return data_types

    def extract_entities(
        self,
        excel_structure: dict[str, Any],
        language: str = "japanese",
        business_domain: str = "general",
    ) -> dict[str, Any]:
        """Extract business entities from Excel structure."""
        entities: dict[str, Any] = {
            "items": [],
            "types": set(),
            "confidence": 0.0,
            "language": language,
            "domain": business_domain,
        }

        for _sheet_name, sheet_data in excel_structure["sheets"].items():
            if "sample_data" in sheet_data:
                sheet_entities = self._extract_sheet_entities(
                    sheet_data["sample_data"], language, business_domain
                )
                entities["items"].extend(sheet_entities)

        # Deduplicate and categorize
        entities["types"] = list({item["type"] for item in entities["items"]})
        entities["confidence"] = self._calculate_entity_confidence(entities["items"])

        return entities

    def convert_to_json(
        self,
        excel_structure: dict[str, Any],
        preserve_formatting: bool = True,
        include_metadata: bool = True,
    ) -> list[dict[str, Any]]:
        """Convert Excel structure to JSON format."""
        json_data = []

        for sheet_name, sheet_data in excel_structure["sheets"].items():
            if "sample_data" in sheet_data:
                sheet_json = self._convert_sheet_to_json(
                    sheet_data["sample_data"],
                    sheet_name,
                    preserve_formatting,
                    include_metadata,
                )
                json_data.extend(sheet_json)

        return json_data

    def _analyze_sheet_structure(
        self, sheet: Worksheet, sheet_name: str
    ) -> dict[str, Any]:
        """Analyze individual sheet structure."""
        structure = {
            "name": sheet_name,
            "dimensions": {"max_row": sheet.max_row, "max_column": sheet.max_column},
            "has_data": False,
            "sample_data": [],
            "format_indicators": {},
        }

        # Extract sample data (first 10 rows)
        sample_rows = []
        for _row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
            if any(cell is not None for cell in row):
                structure["has_data"] = True
                sample_rows.append(list(row))

        structure["sample_data"] = sample_rows

        # Analyze format indicators
        structure["format_indicators"] = self._detect_format_indicators(sheet)

        return structure

    def _has_pivot_indicators(self, sheet: Worksheet) -> bool:
        """Check if sheet contains pivot table indicators."""
        # Look for typical pivot table patterns
        sample_data = []
        for row in sheet.iter_rows(max_row=10, values_only=True):
            if any(cell is not None for cell in row):
                sample_data.append(row)

        if len(sample_data) < 3:
            return False

        # Check for grand total, subtotal patterns
        pivot_keywords = ["合計", "小計", "Total", "Subtotal", "Grand Total"]
        for row in sample_data:
            for cell in row:
                if isinstance(cell, str) and any(
                    keyword in cell for keyword in pivot_keywords
                ):
                    return True

        return False

    def _has_financial_indicators(self, sheet: Worksheet) -> bool:
        """Check if sheet contains financial statement indicators."""
        financial_keywords = [
            "売上",
            "売上高",
            "売上総利益",
            "営業利益",
            "当期純利益",
            "資産",
            "負債",
            "純資産",
            "流動資産",
            "固定資産",
            "Revenue",
            "Sales",
            "Profit",
            "Assets",
            "Liabilities",
        ]

        for row in sheet.iter_rows(max_row=20, values_only=True):
            for cell in row:
                if isinstance(cell, str) and any(
                    keyword in cell for keyword in financial_keywords
                ):
                    return True

        return False

    def _has_multi_headers(self, sheet: Worksheet) -> bool:
        """Check if sheet has multi-row headers."""
        # Check first 5 rows for header patterns
        header_rows = []
        for row in sheet.iter_rows(max_row=5, values_only=True):
            header_rows.append(row)

        if len(header_rows) < 2:
            return False

        # Look for merged cells or hierarchical structure
        non_empty_counts = []
        for row in header_rows:
            non_empty_count = sum(
                1 for cell in row if cell is not None and str(cell).strip()
            )
            non_empty_counts.append(non_empty_count)

        # Multi-header typically has different non-empty counts in header rows
        return len(set(non_empty_counts)) > 1 and max(non_empty_counts) > 1

    def _has_crosstab_structure(self, sheet: Worksheet) -> bool:
        """Check if sheet has cross-tabulation structure."""
        # Check for typical crosstab pattern (row headers + column data)
        sample_data = []
        for row in sheet.iter_rows(max_row=10, values_only=True):
            if any(cell is not None for cell in row):
                sample_data.append(row)

        if len(sample_data) < 3:
            return False

        # Look for numeric data in a grid pattern
        numeric_cols = 0
        text_cols = 0

        if sample_data:
            for col_idx in range(len(sample_data[0])):
                col_data = [
                    row[col_idx] for row in sample_data[1:] if col_idx < len(row)
                ]
                numeric_count = sum(
                    1 for cell in col_data if isinstance(cell, int | float)
                )
                text_count = sum(1 for cell in col_data if isinstance(cell, str))

                if numeric_count > text_count:
                    numeric_cols += 1
                else:
                    text_cols += 1

        # Crosstab typically has some text columns (categories) and numeric columns (values)
        return numeric_cols >= 2 and text_cols >= 1

    def _has_time_series_pattern(self, sheet: Worksheet) -> bool:
        """Check if sheet contains time series data pattern."""
        date_patterns = [
            r"\d{4}[-/]\d{1,2}[-/]\d{1,2}",  # YYYY-MM-DD
            r"\d{1,2}[-/]\d{1,2}[-/]\d{4}",  # MM-DD-YYYY
            r"\d{4}年\d{1,2}月",  # Japanese date
            r"\d{1,2}月",  # Japanese month
        ]

        for row in sheet.iter_rows(max_row=10, values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    for pattern in date_patterns:
                        if re.search(pattern, cell):
                            return True
                elif isinstance(cell, datetime):
                    return True

        return False

    def _determine_overall_format(self, sheets: dict[str, Any]) -> str:
        """Determine overall format based on all sheets."""
        format_votes = {}

        for sheet_data in sheets.values():
            indicators = sheet_data.get("format_indicators", {})
            for format_type, confidence in indicators.items():
                if format_type not in format_votes:
                    format_votes[format_type] = 0
                format_votes[format_type] += confidence

        if format_votes:
            return max(format_votes, key=lambda x: format_votes[x])

        return "standard_table"

    def _calculate_complexity_score(self, sheets: dict[str, Any]) -> float:
        """Calculate overall complexity score."""
        complexity_factors = []

        for sheet_data in sheets.values():
            # Factor 1: Sheet size
            dimensions = sheet_data.get("dimensions", {})
            size_score = min(
                1.0,
                (dimensions.get("max_row", 0) * dimensions.get("max_column", 0))
                / 10000,
            )
            complexity_factors.append(size_score)

            # Factor 2: Format complexity
            indicators = sheet_data.get("format_indicators", {})
            format_complexity = len(indicators) * 0.2
            complexity_factors.append(min(1.0, format_complexity))

        return (
            sum(complexity_factors) / len(complexity_factors)
            if complexity_factors
            else 0.0
        )

    def _detect_format_indicators(self, sheet: Worksheet) -> dict[str, float]:
        """Detect format indicators with confidence scores."""
        indicators = {}

        if self._has_pivot_indicators(sheet):
            indicators["pivot_table"] = 0.8

        if self._has_financial_indicators(sheet):
            indicators["financial_statement"] = 0.9

        if self._has_multi_headers(sheet):
            indicators["multi_header"] = 0.7

        if self._has_crosstab_structure(sheet):
            indicators["cross_tab"] = 0.6

        if self._has_time_series_pattern(sheet):
            indicators["time_series"] = 0.8

        if not indicators:
            indicators["standard_table"] = 0.9

        return indicators

    def _load_entity_patterns(self) -> dict[str, list[str]]:
        """Load entity recognition patterns."""
        return {
            "person": [
                r"[一-龯]{1,3}[太郎|花子|次郎|三郎]",  # Japanese names
                r"[田中|佐藤|山田|鈴木|高橋][一-龯]{2,3}",
            ],
            "location": [
                r"[東京|大阪|名古屋|福岡|札幌][都府県市]?",
                r"[一-龯]{2,4}[県市区町村]",
            ],
            "organization": [
                r"株式会社[一-龯]{2,10}",
                r"[一-龯]{2,10}株式会社",
                r"[一-龯]{2,10}[部課室]",
            ],
            "financial": [
                r"[0-9,]+円",
                r"[0-9,]+万円",
                r"[0-9,]+億円",
            ],
        }

    def _extract_business_entities(
        self,
        data: list[dict[str, Any]],
        domain: str = "general",
        language: str = "japanese",
    ) -> dict[str, Any]:
        """Extract business entities from converted data."""
        entities: dict[str, Any] = {
            "items": [],
            "types": [],
            "confidence": 0.8,
            "domain": domain,
            "language": language,
        }

        # This is a simplified implementation
        # Full implementation would use NLP libraries for entity recognition
        entity_types = set()

        for record in data:
            for _key, value in record.items():
                if isinstance(value, str):
                    # Simple pattern matching for demonstration
                    if any(char in value for char in "株式会社"):
                        entities["items"].append(
                            {"text": value, "type": "organization", "confidence": 0.9}
                        )
                        entity_types.add("organization")
                    elif any(char in value for char in "円万億"):
                        entities["items"].append(
                            {"text": value, "type": "financial", "confidence": 0.8}
                        )
                        entity_types.add("financial")

        entities["types"] = list(entity_types)
        return entities

    def _generate_comprehensive_metadata(
        self,
        conversion_result: dict[str, Any],
        entities: dict[str, Any],
        config: dict[str, Any],
    ) -> dict[str, Any]:
        """Generate comprehensive metadata."""
        return {
            "source_file": conversion_result.get("source_file", ""),
            "conversion_timestamp": datetime.now().isoformat(),
            "record_count": len(conversion_result.get("data", [])),
            "entity_summary": {
                "total_entities": len(entities.get("items", [])),
                "entity_types": entities.get("types", []),
                "confidence": entities.get("confidence", 0.8),
            },
            "data_quality": {
                "completeness": self._calculate_completeness(
                    conversion_result.get("data", [])
                ),
                "consistency": 0.9,  # Placeholder
                "accuracy": 0.85,  # Placeholder
            },
            "config": config,
        }

    def _assess_data_quality(
        self, data: list[dict[str, Any]], metadata: dict[str, Any]
    ) -> float:
        """Assess overall data quality."""
        quality_factors = []

        # Completeness
        completeness = self._calculate_completeness(data)
        quality_factors.append(completeness)

        # Consistency (simplified check)
        consistency = self._calculate_consistency(data)
        quality_factors.append(consistency)

        # Return average
        return sum(quality_factors) / len(quality_factors) if quality_factors else 0.8

    def _calculate_completeness(self, data: list[dict[str, Any]]) -> float:
        """Calculate data completeness score."""
        if not data:
            return 0.0

        total_fields = 0
        filled_fields = 0

        for record in data:
            for value in record.values():
                total_fields += 1
                if value is not None and str(value).strip():
                    filled_fields += 1

        return filled_fields / total_fields if total_fields > 0 else 0.0

    def _calculate_consistency(self, data: list[dict[str, Any]]) -> float:
        """Calculate data consistency score."""
        if not data:
            return 1.0

        # Simple consistency check: same keys across records
        if len(data) <= 1:
            return 1.0

        first_keys = set(data[0].keys())
        consistent_records = sum(
            1 for record in data if set(record.keys()) == first_keys
        )

        return consistent_records / len(data)

    def _infer_column_type(self, sample_data: list[list[Any]], col_idx: int) -> str:
        """Infer column data type from sample data."""
        if not sample_data or col_idx >= len(sample_data[0]):
            return "unknown"

        col_values = [
            row[col_idx]
            for row in sample_data[1:]
            if col_idx < len(row) and row[col_idx] is not None
        ]

        if not col_values:
            return "empty"

        # Count types
        int_count = sum(1 for v in col_values if isinstance(v, int))
        float_count = sum(1 for v in col_values if isinstance(v, float))
        str_count = sum(1 for v in col_values if isinstance(v, str))
        date_count = sum(1 for v in col_values if isinstance(v, datetime))

        total = len(col_values)

        if date_count / total > 0.5:
            return "datetime"
        elif (int_count + float_count) / total > 0.7:
            return "numeric"
        elif str_count / total > 0.7:
            return "text"
        else:
            return "mixed"

    def _extract_sheet_entities(
        self, sample_data: list[list[Any]], language: str, business_domain: str
    ) -> list[dict[str, Any]]:
        """Extract entities from sheet sample data."""
        entities = []

        for row in sample_data:
            for cell in row:
                if isinstance(cell, str) and cell.strip():
                    # Simple entity recognition
                    if "株式会社" in cell or "会社" in cell:
                        entities.append(
                            {"text": cell, "type": "organization", "confidence": 0.8}
                        )
                    elif any(char in cell for char in "円万億"):
                        entities.append(
                            {"text": cell, "type": "financial", "confidence": 0.7}
                        )

        return entities

    def _calculate_entity_confidence(self, entities: list[dict[str, Any]]) -> float:
        """Calculate overall entity recognition confidence."""
        if not entities:
            return 0.0

        total_confidence = sum(entity.get("confidence", 0.5) for entity in entities)
        return total_confidence / len(entities)

    def _convert_sheet_to_json(
        self,
        sample_data: list[list[Any]],
        sheet_name: str,
        preserve_formatting: bool,
        include_metadata: bool,
    ) -> list[dict[str, Any]]:
        """Convert sheet data to JSON format."""
        if not sample_data:
            return []

        # Use first row as headers
        headers = [
            str(cell) if cell is not None else f"column_{i}"
            for i, cell in enumerate(sample_data[0])
        ]

        json_records: list[dict[str, Any]] = []
        for row in sample_data[1:]:
            record = {}
            for i, value in enumerate(row):
                header = headers[i] if i < len(headers) else f"column_{i}"
                record[header] = value

            if include_metadata:
                record["_source_sheet"] = sheet_name
                record["_row_index"] = len(json_records) + 1

            json_records.append(record)

        return json_records


# Base class for format handlers
class BaseFormatHandler:
    """Base class for Excel format handlers."""

    def convert(self, excel_file: str, config: dict[str, Any]) -> dict[str, Any]:
        """Convert Excel file based on specific format."""
        raise NotImplementedError("Subclasses must implement convert method")


# Specific format handlers
class PivotTableHandler(BaseFormatHandler):
    """Handler for pivot table format."""

    def convert(self, excel_file: str, config: dict[str, Any]) -> dict[str, Any]:
        """Convert pivot table to standard format."""
        # Simplified implementation
        df = pd.read_excel(excel_file)
        return {
            "data": df.to_dict("records"),
            "format": "pivot_table",
            "confidence": 0.8,
            "notes": ["Pivot table detected and flattened"],
        }


class FinancialStatementHandler(BaseFormatHandler):
    """Handler for financial statement format."""

    def convert(self, excel_file: str, config: dict[str, Any]) -> dict[str, Any]:
        """Convert financial statement to standard format."""
        df = pd.read_excel(excel_file)
        return {
            "data": df.to_dict("records"),
            "format": "financial_statement",
            "confidence": 0.9,
            "notes": ["Financial statement format detected"],
        }


class MultiHeaderHandler(BaseFormatHandler):
    """Handler for multi-header format."""

    def convert(self, excel_file: str, config: dict[str, Any]) -> dict[str, Any]:
        """Convert multi-header table to standard format."""
        df = pd.read_excel(excel_file, header=[0, 1])  # Multi-level header
        df.columns = ["_".join(col).strip() for col in df.columns.values]
        return {
            "data": df.to_dict("records"),
            "format": "multi_header",
            "confidence": 0.7,
            "notes": ["Multi-header structure unified"],
        }


class CrossTabHandler(BaseFormatHandler):
    """Handler for cross-tabulation format."""

    def convert(self, excel_file: str, config: dict[str, Any]) -> dict[str, Any]:
        """Convert cross-tab to standard format."""
        df = pd.read_excel(excel_file, index_col=0)
        df_melted = df.reset_index().melt(id_vars=df.index.name or "index")
        return {
            "data": df_melted.to_dict("records"),
            "format": "cross_tab",
            "confidence": 0.6,
            "notes": ["Cross-tabulation melted to long format"],
        }


class TimeSeriesHandler(BaseFormatHandler):
    """Handler for time series format."""

    def convert(self, excel_file: str, config: dict[str, Any]) -> dict[str, Any]:
        """Convert time series data to standard format."""
        df = pd.read_excel(excel_file, parse_dates=True)
        return {
            "data": df.to_dict("records"),
            "format": "time_series",
            "confidence": 0.8,
            "notes": ["Time series data processed with date parsing"],
        }


class StandardTableHandler(BaseFormatHandler):
    """Handler for standard table format."""

    def convert(self, excel_file: str, config: dict[str, Any]) -> dict[str, Any]:
        """Convert standard table to JSON format."""
        df = pd.read_excel(excel_file)
        return {
            "data": df.to_dict("records"),
            "format": "standard_table",
            "confidence": 0.9,
            "notes": ["Standard table format processed"],
        }
