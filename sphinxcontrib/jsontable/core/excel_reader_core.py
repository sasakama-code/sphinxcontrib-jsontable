"""Excel Reader Core - Production implementation of Excel reading operations.

Simplified, focused implementation of Excel file reading with essential
functionality and clean error handling.

CLAUDE.md Code Excellence Compliance:
- Single Responsibility: Excel file reading operations only
- DRY Principle: Centralized reading logic without duplication
- SOLID Principles: Interface implementation with dependency injection
"""

import os
from pathlib import Path
from typing import List, Optional, Union

import pandas as pd
from openpyxl import load_workbook

from ..errors.excel_errors import (
    ExcelFileNotFoundError,
    ExcelProcessingError,
    FileAccessError,
    SecurityValidationError,
    WorksheetNotFoundError,
)
from .excel_reader_interface import IExcelReader
from .excel_workbook_info import ReadResult, WorkbookInfo


class ExcelReader(IExcelReader):
    """Production implementation of Excel file reading functionality.

    Provides essential Excel reading capabilities with security validation
    and comprehensive error handling. Focuses on core functionality without
    unnecessary complexity.
    """

    def __init__(
        self,
        max_file_size: int = 100 * 1024 * 1024,  # 100MB
        allowed_extensions: Optional[List[str]] = None,
        enable_security_validation: bool = True,
    ):
        """Initialize Excel reader with configuration.

        Args:
            max_file_size: Maximum allowed file size in bytes
            allowed_extensions: List of allowed file extensions
            enable_security_validation: Whether to perform security validation
        """
        self.max_file_size = max_file_size
        self.allowed_extensions = allowed_extensions or [
            ".xlsx",
            ".xls",
            ".xlsm",
            ".xltm",
        ]
        self.enable_security_validation = enable_security_validation

    def validate_file(self, file_path: Union[str, Path]) -> WorkbookInfo:
        """Validate Excel file and return workbook information.

        Args:
            file_path: Path to Excel file

        Returns:
            WorkbookInfo containing file metadata
        """
        file_path = Path(file_path)

        try:
            # Basic validation
            self._validate_file_existence(file_path)
            self._validate_file_extension(file_path)
            self._validate_file_size(file_path)

            # Load workbook for detailed inspection
            workbook = load_workbook(file_path, data_only=True)
            sheet_names = workbook.sheetnames

            # Security inspection
            has_macros = self._check_macros(file_path)
            has_external_links = self._check_external_links(workbook)

            return WorkbookInfo(
                file_path=file_path,
                sheet_names=sheet_names,
                has_macros=has_macros,
                has_external_links=has_external_links,
                file_size=file_path.stat().st_size,
                format_type=file_path.suffix,
            )

        except Exception as e:
            if isinstance(
                e, (ExcelFileNotFoundError, FileAccessError, SecurityValidationError)
            ):
                raise
            raise ExcelProcessingError(f"Failed to validate Excel file: {e}") from e

    def get_sheet_names(self, file_path: Union[str, Path]) -> List[str]:
        """Get list of sheet names from Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of sheet names
        """
        try:
            workbook = load_workbook(file_path, data_only=True)
            return workbook.sheetnames
        except Exception as e:
            raise ExcelProcessingError(f"Failed to read sheet names: {e}") from e

    def read_workbook(
        self,
        file_path: Union[str, Path],
        sheet_name: Optional[str] = None,
        sheet_index: Optional[int] = None,
        **kwargs,
    ) -> ReadResult:
        """Read Excel workbook and return data with metadata.

        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet name to read
            sheet_index: Specific sheet index to read (0-based)
            **kwargs: Additional pandas read_excel options

        Returns:
            ReadResult containing DataFrame and metadata
        """
        try:
            # Validate file first
            workbook_info = self.validate_file(file_path)

            # Determine target sheet
            target_sheet = self._resolve_target_sheet(
                workbook_info.sheet_names, sheet_name, sheet_index
            )

            # Read data using pandas
            dataframe = pd.read_excel(file_path, sheet_name=target_sheet, **kwargs)

            # Create metadata
            metadata = {
                "sheet_name": target_sheet,
                "original_shape": dataframe.shape,
                "read_options": kwargs,
            }

            return ReadResult(
                dataframe=dataframe,
                workbook_info=workbook_info,
                metadata=metadata,
            )

        except Exception as e:
            if isinstance(e, (ExcelFileNotFoundError, WorksheetNotFoundError)):
                raise
            raise ExcelProcessingError(f"Failed to read Excel workbook: {e}") from e

    def read_sheet(
        self, file_path: Union[str, Path], sheet_identifier: Union[str, int], **kwargs
    ) -> ReadResult:
        """Read specific sheet from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_identifier: Sheet name (str) or index (int)
            **kwargs: Additional reading options

        Returns:
            ReadResult containing sheet data and metadata
        """
        if isinstance(sheet_identifier, str):
            return self.read_workbook(file_path, sheet_name=sheet_identifier, **kwargs)
        else:
            return self.read_workbook(file_path, sheet_index=sheet_identifier, **kwargs)

    # Private helper methods

    def _validate_file_existence(self, file_path: Path) -> None:
        """Validate file existence and accessibility."""
        if not file_path.exists():
            raise ExcelFileNotFoundError(f"File not found: {file_path}")

        if not file_path.is_file():
            raise FileAccessError(f"Path is not a file: {file_path}")

        if not os.access(file_path, os.R_OK):
            raise FileAccessError(f"File is not readable: {file_path}")

    def _validate_file_extension(self, file_path: Path) -> None:
        """Validate file extension is allowed."""
        if file_path.suffix.lower() not in self.allowed_extensions:
            raise ExcelProcessingError(
                f"Unsupported file extension: {file_path.suffix}. "
                f"Allowed: {self.allowed_extensions}"
            )

    def _validate_file_size(self, file_path: Path) -> None:
        """Validate file size is within limits."""
        file_size = file_path.stat().st_size
        if file_size > self.max_file_size:
            raise FileAccessError(
                f"File too large: {file_size} bytes. "
                f"Maximum allowed: {self.max_file_size} bytes"
            )

    def _check_macros(self, file_path: Path) -> bool:
        """Check if file contains macros."""
        # Simple check based on file extension
        macro_extensions = {".xlsm", ".xltm", ".xlam"}
        return file_path.suffix.lower() in macro_extensions

    def _check_external_links(self, workbook) -> bool:
        """Check if workbook contains external links."""
        try:
            # Simple check for external links in defined names
            return bool(workbook.defined_names)
        except Exception:
            return False

    def _resolve_target_sheet(
        self,
        available_sheets: List[str],
        sheet_name: Optional[str],
        sheet_index: Optional[int],
    ) -> str:
        """Resolve target sheet name from parameters."""
        if sheet_name:
            if sheet_name not in available_sheets:
                raise WorksheetNotFoundError(
                    f"Sheet '{sheet_name}' not found. Available: {available_sheets}"
                )
            return sheet_name

        if sheet_index is not None:
            if not 0 <= sheet_index < len(available_sheets):
                raise WorksheetNotFoundError(
                    f"Sheet index {sheet_index} out of range. "
                    f"Available indices: 0-{len(available_sheets) - 1}"
                )
            return available_sheets[sheet_index]

        # Default to first sheet
        return available_sheets[0] if available_sheets else "Sheet1"
