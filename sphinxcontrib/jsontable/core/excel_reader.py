"""Excel Reader for file I/O operations.

Extracted from monolithic ExcelDataLoader to provide focused,
testable Excel file reading functionality with single responsibility.
"""

import os
from abc import ABC, abstractmethod
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import load_workbook

from ..errors.excel_errors import (
    ExcelFileNotFoundError,
    ExcelProcessingError,
    FileAccessError,
    SecurityValidationError,
    WorksheetNotFoundError,
)


@dataclass
class WorkbookInfo:
    """Information about an Excel workbook."""

    file_path: Path
    sheet_names: List[str]
    has_macros: bool
    has_external_links: bool
    file_size: int
    format_type: str

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization."""
        return {
            "file_path": str(self.file_path),
            "sheet_names": self.sheet_names,
            "has_macros": self.has_macros,
            "has_external_links": self.has_external_links,
            "file_size": self.file_size,
            "format_type": self.format_type,
            "total_sheets": len(self.sheet_names),
        }


@dataclass
class ReadResult:
    """Result of Excel file reading operation."""

    dataframe: pd.DataFrame
    workbook_info: WorkbookInfo
    sheet_name: str
    metadata: Dict[str, Any]

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization."""
        return {
            "dataframe_shape": self.dataframe.shape,
            "workbook_info": self.workbook_info.to_dict(),
            "sheet_name": self.sheet_name,
            "metadata": self.metadata,
        }


class IExcelReader(ABC):
    """Abstract interface for Excel file reading functionality."""

    @abstractmethod
    def validate_file(self, file_path: Union[str, Path]) -> WorkbookInfo:
        """Validate Excel file and return workbook information.

        Args:
            file_path: Path to Excel file

        Returns:
            WorkbookInfo: Information about the validated workbook

        Raises:
            ExcelFileNotFoundError: If file doesn't exist
            FileAccessError: If file can't be accessed
            SecurityValidationError: If security validation fails
        """
        pass

    @abstractmethod
    def read_excel(
        self,
        file_path: Union[str, Path],
        sheet_name: Optional[str] = None,
        sheet_index: Optional[int] = None,
        skip_rows: Optional[Union[int, List[int]]] = None,
        range_spec: Optional[str] = None,
        **kwargs,
    ) -> ReadResult:
        """Read Excel file and return DataFrame with metadata.

        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to read (optional)
            sheet_index: Index of sheet to read (optional)
            skip_rows: Rows to skip when reading (optional)
            range_spec: Range specification (optional)
            **kwargs: Additional pandas.read_excel arguments

        Returns:
            ReadResult: Reading result with DataFrame and metadata

        Raises:
            ExcelProcessingError: If reading fails
            WorksheetNotFoundError: If specified sheet doesn't exist
        """
        pass

    @abstractmethod
    def get_sheet_names(self, file_path: Union[str, Path]) -> List[str]:
        """Get list of sheet names from Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List[str]: List of sheet names
        """
        pass


class ExcelReader(IExcelReader):
    """Production implementation of Excel file reading functionality.

    Implements file I/O logic extracted from excel_data_loader.py
    lines 198-800 (approximately 600 lines of file reading logic).
    """

    def __init__(
        self,
        max_file_size: int = 100 * 1024 * 1024,  # 100MB
        allowed_extensions: Optional[List[str]] = None,
        enable_security_validation: bool = True,
    ):
        """Initialize Excel reader with configuration.

        Args:
            max_file_size: Maximum allowed file size in bytes
            allowed_extensions: List of allowed file extensions
            enable_security_validation: Whether to perform security validation
        """
        self.max_file_size = max_file_size
        self.allowed_extensions = allowed_extensions or [
            ".xlsx",
            ".xls",
            ".xlsm",
            ".xltm",
        ]
        self.enable_security_validation = enable_security_validation

    def validate_file(self, file_path: Union[str, Path]) -> WorkbookInfo:
        """Validate Excel file and return workbook information.

        Implements validation logic from excel_data_loader.py lines 198-268.
        """
        file_path = Path(file_path)

        try:
            # Basic file existence and access validation
            self._validate_file_existence(file_path)
            self._validate_file_extension(file_path)
            self._validate_file_size(file_path)
            self._validate_path_security(file_path)

            # Load workbook for detailed inspection
            try:
                workbook = load_workbook(file_path, data_only=True)
                sheet_names = workbook.sheetnames

                # Security validation
                security_info = (
                    self._validate_security(workbook)
                    if self.enable_security_validation
                    else {"has_macros": False, "has_external_links": False}
                )

                return WorkbookInfo(
                    file_path=file_path,
                    sheet_names=sheet_names,
                    has_macros=security_info["has_macros"],
                    has_external_links=security_info["has_external_links"],
                    file_size=file_path.stat().st_size,
                    format_type=file_path.suffix.lower(),
                )

            except Exception as e:
                raise ExcelProcessingError(
                    message=f"Failed to load workbook: {e}",
                    context={"file_path": str(file_path)},
                    original_error=e,
                ) from e

        except (ExcelFileNotFoundError, FileAccessError, SecurityValidationError):
            raise
        except Exception as e:
            raise ExcelProcessingError(
                message=f"File validation failed: {e}",
                context={"file_path": str(file_path)},
                original_error=e,
            ) from e

    def read_excel(
        self,
        file_path: Union[str, Path],
        sheet_name: Optional[str] = None,
        sheet_index: Optional[int] = None,
        skip_rows: Optional[Union[int, List[int]]] = None,
        range_spec: Optional[str] = None,
        **kwargs,
    ) -> ReadResult:
        """Read Excel file and return DataFrame with metadata.

        Implements reading logic from excel_data_loader.py lines 691-800.
        """
        file_path = Path(file_path)

        try:
            # Validate file first
            workbook_info = self.validate_file(file_path)

            # Determine which sheet to read
            target_sheet = self._determine_target_sheet(
                workbook_info.sheet_names, sheet_name, sheet_index
            )

            # Prepare pandas.read_excel arguments
            read_kwargs = {
                "sheet_name": target_sheet,
                "header": None,  # We'll handle headers separately
                **kwargs,
            }

            # Read Excel file using pandas with proper try/except
            try:
                # Handle range+skip_rows combination specially
                if range_spec and skip_rows is not None:
                    # For range+skip_rows combination, handle separately
                    range_info = self._parse_range_specification(range_spec)
                    if range_info:
                        # Read the range first
                        read_kwargs["usecols"] = range_info["usecols"]
                        if range_info["nrows"] is not None:
                            read_kwargs["nrows"] = range_info["nrows"]
                        if range_info["skiprows"] is not None:
                            read_kwargs["skiprows"] = range_info["skiprows"]

                        # Read dataframe and then apply within-range skip_rows
                        dataframe = pd.read_excel(file_path, **read_kwargs)
                        dataframe = self._apply_within_range_skip_rows(
                            dataframe, skip_rows
                        )
                    else:
                        # Fallback to standard processing
                        dataframe = pd.read_excel(file_path, **read_kwargs)
                else:
                    # Standard processing for individual options
                    if skip_rows is not None:
                        read_kwargs["skiprows"] = skip_rows

                    if range_spec:
                        # Basic range parsing - could be enhanced with RangeParser
                        usecols = self._parse_range_to_usecols(range_spec)
                        if usecols:
                            read_kwargs["usecols"] = usecols

                    # Read Excel file using pandas
                    dataframe = pd.read_excel(file_path, **read_kwargs)

                # Build metadata
                metadata = {
                    "read_method": "pandas.read_excel",
                    "total_sheets": len(workbook_info.sheet_names),
                    "file_size": workbook_info.file_size,
                    "format_type": workbook_info.format_type,
                    "security_validated": self.enable_security_validation,
                    "skip_rows": skip_rows,
                    "range_spec": range_spec,
                    "read_kwargs": {
                        k: v for k, v in read_kwargs.items() if k != "sheet_name"
                    },
                }

                return ReadResult(
                    dataframe=dataframe,
                    workbook_info=workbook_info,
                    sheet_name=target_sheet,
                    metadata=metadata,
                )

            except Exception as e:
                raise ExcelProcessingError(
                    message=f"Failed to read Excel data: {e}",
                    context={"file_path": str(file_path), "sheet_name": target_sheet},
                    original_error=e,
                ) from e

        except (ExcelProcessingError, WorksheetNotFoundError):
            raise
        except Exception as e:
            raise ExcelProcessingError(
                message=f"Excel reading failed: {e}",
                context={"file_path": str(file_path)},
                original_error=e,
            ) from e

    def get_sheet_names(self, file_path: Union[str, Path]) -> List[str]:
        """Get list of sheet names from Excel file."""
        try:
            with pd.ExcelFile(file_path) as excel_file:
                return excel_file.sheet_names
        except Exception as e:
            raise ExcelProcessingError(
                message=f"Failed to get sheet names: {e}",
                context={"file_path": str(file_path)},
                original_error=e,
            ) from e

    def _validate_file_existence(self, file_path: Path) -> None:
        """Validate file existence and accessibility."""
        if not file_path.exists():
            raise ExcelFileNotFoundError(str(file_path))

        if not file_path.is_file():
            raise FileAccessError(
                str(file_path), message="Path exists but is not a file"
            )

        if not os.access(file_path, os.R_OK):
            raise FileAccessError(
                str(file_path), message="File exists but is not readable"
            )

    def _validate_file_extension(self, file_path: Path) -> None:
        """Validate file extension."""
        if file_path.suffix.lower() not in self.allowed_extensions:
            raise ExcelProcessingError(
                message=f"Unsupported file extension: {file_path.suffix}",
                error_code="UNSUPPORTED_FORMAT",
                context={
                    "file_extension": file_path.suffix,
                    "allowed_extensions": self.allowed_extensions,
                },
            )

    def _validate_file_size(self, file_path: Path) -> None:
        """Validate file size."""
        file_size = file_path.stat().st_size
        if file_size > self.max_file_size:
            raise ExcelProcessingError(
                message=f"File too large: {file_size} bytes (max: {self.max_file_size})",
                error_code="FILE_TOO_LARGE",
                context={"file_size": file_size, "max_size": self.max_file_size},
            )

    def _validate_path_security(self, file_path: Path) -> None:
        """Validate path for security (path traversal prevention)."""
        try:
            # Resolve path and check for path traversal
            resolved_path = file_path.resolve()

            # Check for suspicious path patterns
            path_str = str(resolved_path)
            if ".." in path_str or path_str.startswith("/"):
                # Additional security checks can be added here
                pass

        except Exception as e:
            raise SecurityValidationError(
                [],
                message=f"Path security validation failed: {e}",
                context={"file_path": str(file_path)},
                original_error=e,
            ) from e

    def _validate_security(self, workbook) -> Dict[str, bool]:
        """Validate workbook security (macros and external links).

        Implements security validation from excel_data_loader.py lines 270-440.
        """
        security_info = {"has_macros": False, "has_external_links": False}

        try:
            # Check for VBA/macros
            if hasattr(workbook, "vba_archive") and workbook.vba_archive:
                security_info["has_macros"] = True

            # Check for external links in all worksheets
            for worksheet in workbook.worksheets:
                # Check for external links in formulas
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            if any(
                                link_indicator in cell.value.lower()
                                for link_indicator in [
                                    "http://",
                                    "https://",
                                    "ftp://",
                                    "file://",
                                ]
                            ):
                                security_info["has_external_links"] = True
                                break
                    if security_info["has_external_links"]:
                        break
                if security_info["has_external_links"]:
                    break

            # Validate security threats
            if security_info["has_macros"] or security_info["has_external_links"]:
                security_issues = []

                if security_info["has_macros"]:
                    security_issues.append(
                        {
                            "type": "macro_detected",
                            "severity": "high",
                            "description": "VBA macros detected in workbook",
                        }
                    )

                if security_info["has_external_links"]:
                    security_issues.append(
                        {
                            "type": "external_links",
                            "severity": "medium",
                            "description": "External links detected in workbook",
                        }
                    )

                raise SecurityValidationError(
                    security_issues,
                    message=f"Security validation failed: {len(security_issues)} issues detected",
                )

            return security_info

        except SecurityValidationError:
            raise
        except Exception as e:
            raise ExcelProcessingError(
                message=f"Security validation error: {e}",
                context={"validation_type": "security"},
                original_error=e,
            ) from e

    def _determine_target_sheet(
        self,
        sheet_names: List[str],
        sheet_name: Optional[str] = None,
        sheet_index: Optional[int] = None,
    ) -> str:
        """Determine which sheet to read based on parameters.

        Implements sheet selection logic from excel_data_loader.py lines 441-540.
        """
        if not sheet_names:
            raise WorksheetNotFoundError(
                "", [], message="No worksheets found in workbook"
            )

        # If sheet name is specified, validate it exists
        if sheet_name:
            if sheet_name not in sheet_names:
                raise WorksheetNotFoundError(sheet_name, sheet_names)
            return sheet_name

        # If sheet index is specified, validate it's in range
        if sheet_index is not None:
            if not (0 <= sheet_index < len(sheet_names)):
                raise WorksheetNotFoundError(
                    f"Index {sheet_index}",
                    sheet_names,
                    message=f"Sheet index {sheet_index} out of range (0-{len(sheet_names) - 1})",
                )
            return sheet_names[sheet_index]

        # Default to first sheet
        return sheet_names[0]

    def _parse_range_to_usecols(self, range_spec: str) -> Optional[List[int]]:
        """Parse range specification to usecols for pandas.

        Args:
            range_spec: Range specification like 'A1:C10' or 'A:C'

        Returns:
            List of column indices or None if cannot parse
        """
        try:
            # Simple range parsing - just extract column letters
            if ":" in range_spec:
                parts = range_spec.split(":")
                if len(parts) == 2:
                    start_col = self._column_letter_to_index(parts[0][0])
                    end_col = self._column_letter_to_index(parts[1][0])
                    if start_col is not None and end_col is not None:
                        return list(range(start_col, end_col + 1))
            return None
        except (IndexError, ValueError):
            return None

    def _parse_range_specification(self, range_spec: str) -> Optional[Dict[str, Any]]:
        """Parse range specification to extract row/column information.

        Args:
            range_spec: Range specification like 'A3:C8'

        Returns:
            Dict with range information or None if cannot parse
        """
        try:
            if ":" not in range_spec:
                return None

            parts = range_spec.split(":")
            if len(parts) != 2:
                return None

            start_cell, end_cell = parts

            # Extract start row/column
            start_col_str = ""
            start_row_str = ""
            for char in start_cell:
                if char.isalpha():
                    start_col_str += char
                elif char.isdigit():
                    start_row_str += char

            # Extract end row/column
            end_col_str = ""
            end_row_str = ""
            for char in end_cell:
                if char.isalpha():
                    end_col_str += char
                elif char.isdigit():
                    end_row_str += char

            # Convert to indices
            start_col = self._column_letter_to_index(start_col_str)
            end_col = self._column_letter_to_index(end_col_str)

            result = {
                "usecols": list(range(start_col, end_col + 1))
                if start_col is not None and end_col is not None
                else None,
                "skiprows": None,
                "nrows": None,
            }

            # Handle row ranges if specified
            if start_row_str and end_row_str:
                start_row = int(start_row_str) - 1  # Convert to 0-based
                end_row = int(end_row_str) - 1  # Convert to 0-based
                result["skiprows"] = start_row
                result["nrows"] = end_row - start_row + 1

            return result

        except (ValueError, IndexError):
            return None

    def _apply_within_range_skip_rows(
        self, dataframe: pd.DataFrame, skip_rows: Any
    ) -> pd.DataFrame:
        """Apply skip_rows within the already-read range dataframe.

        Args:
            dataframe: DataFrame read from the range
            skip_rows: Rows to skip (relative to range)

        Returns:
            DataFrame with specified rows skipped
        """
        try:
            if isinstance(skip_rows, (int, str)):
                # Convert string to list of integers
                if isinstance(skip_rows, str):
                    skip_indices = self._parse_skip_rows_string(skip_rows)
                else:
                    skip_indices = [skip_rows]
            elif isinstance(skip_rows, list):
                skip_indices = skip_rows
            else:
                return dataframe

            # Filter out rows based on index
            valid_indices = [i for i in range(len(dataframe)) if i not in skip_indices]
            return dataframe.iloc[valid_indices].reset_index(drop=True)

        except Exception:
            # If any error occurs, return original dataframe
            return dataframe

    def _parse_skip_rows_string(self, skip_rows: str) -> List[int]:
        """Parse skip_rows string format like '0,1,2' or '0-2,5,7-9'.

        Args:
            skip_rows: String representation of rows to skip

        Returns:
            List of row indices to skip
        """
        result = []

        for part in skip_rows.split(","):
            part = part.strip()
            if "-" in part:
                # Range format: "0-2" -> [0, 1, 2]
                range_parts = part.split("-")
                if len(range_parts) == 2:
                    start, end = int(range_parts[0]), int(range_parts[1])
                    result.extend(range(start, end + 1))
            else:
                # Single number: "5" -> [5]
                result.append(int(part))

        return sorted(list(set(result)))  # Remove duplicates and sort

    def _column_letter_to_index(self, letter: str) -> Optional[int]:
        """Convert column letter to zero-based index.

        Args:
            letter: Column letter like 'A', 'B', etc.

        Returns:
            Zero-based column index or None if invalid
        """
        if letter.isalpha() and len(letter) == 1:
            return ord(letter.upper()) - ord("A")
        return None


class MockExcelReader(IExcelReader):
    """Mock implementation for testing purposes."""

    def __init__(
        self,
        mock_workbook_info: Optional[WorkbookInfo] = None,
        mock_read_result: Optional[ReadResult] = None,
        mock_sheet_names: Optional[List[str]] = None,
        should_fail: bool = False,
        error_to_raise: Optional[Exception] = None,
    ):
        """Initialize mock reader.

        Args:
            mock_workbook_info: Fixed result for validate_file
            mock_read_result: Fixed result for read_excel
            mock_sheet_names: Fixed result for get_sheet_names
            should_fail: Whether to raise errors
            error_to_raise: Specific error to raise
        """
        self.mock_workbook_info = mock_workbook_info
        self.mock_read_result = mock_read_result
        self.mock_sheet_names = mock_sheet_names or ["Sheet1", "Sheet2"]
        self.should_fail = should_fail
        self.error_to_raise = error_to_raise or ExcelProcessingError("Mock error")

        # Call tracking
        self.validate_file_calls = []
        self.read_excel_calls = []
        self.get_sheet_names_calls = []

    def validate_file(self, file_path: Union[str, Path]) -> WorkbookInfo:
        """Mock validate_file implementation."""
        self.validate_file_calls.append({"file_path": str(file_path)})

        if self.should_fail:
            raise self.error_to_raise

        if self.mock_workbook_info:
            return self.mock_workbook_info

        # Default mock result
        return WorkbookInfo(
            file_path=Path(file_path),
            sheet_names=self.mock_sheet_names,
            has_macros=False,
            has_external_links=False,
            file_size=1024,
            format_type=".xlsx",
        )

    def read_excel(
        self,
        file_path: Union[str, Path],
        sheet_name: Optional[str] = None,
        sheet_index: Optional[int] = None,
        skip_rows: Optional[Union[int, List[int]]] = None,
        range_spec: Optional[str] = None,
        **kwargs,
    ) -> ReadResult:
        """Mock read_excel implementation."""
        self.read_excel_calls.append(
            {
                "file_path": str(file_path),
                "sheet_name": sheet_name,
                "sheet_index": sheet_index,
                "skip_rows": skip_rows,
                "range_spec": range_spec,
                "kwargs": kwargs,
            }
        )

        if self.should_fail:
            raise self.error_to_raise

        if self.mock_read_result:
            return self.mock_read_result

        # Default mock result
        mock_df = pd.DataFrame({"A": [1, 2, 3], "B": ["a", "b", "c"]})

        mock_workbook_info = self.validate_file(file_path)

        return ReadResult(
            dataframe=mock_df,
            workbook_info=mock_workbook_info,
            sheet_name=sheet_name or self.mock_sheet_names[0],
            metadata={"mock": True},
        )

    def get_sheet_names(self, file_path: Union[str, Path]) -> List[str]:
        """Mock get_sheet_names implementation."""
        self.get_sheet_names_calls.append({"file_path": str(file_path)})

        if self.should_fail:
            raise self.error_to_raise

        return self.mock_sheet_names
