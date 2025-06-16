"""Excel Data Loader エッジケースカバレッジ向上テスト."""

import shutil
import tempfile
from pathlib import Path
from unittest.mock import Mock, patch

import pandas as pd
import pytest
from openpyxl import Workbook

# Excel対応がある場合のみテストを実行
try:
    from sphinxcontrib.jsontable.excel_data_loader import (
        ExcelDataLoader,
        RangeSpecificationError,
    )

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@pytest.mark.skipif(not EXCEL_AVAILABLE, reason="Excel support not available")
class TestEdgeCaseCoverageImprovement:
    """エッジケースカバレッジ向上テスト."""

    def setup_method(self):
        """各テストメソッドの前に実行される."""
        self.temp_dir = tempfile.mkdtemp()
        self.loader = ExcelDataLoader(self.temp_dir)

    def teardown_method(self):
        """各テストメソッドの後に実行される."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def create_test_excel(self, filename: str, size_mb: int = 1) -> str:
        """指定サイズのテスト用Excelファイルを作成."""
        file_path = str(Path(self.temp_dir) / filename)

        wb = Workbook()
        ws = wb.active

        # 小さなデータを作成
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["A2"] = "Data1"
        ws["B2"] = "Data2"

        wb.save(file_path)
        return file_path

    def create_large_excel_file(self, filename: str) -> str:
        """大きなExcelファイルを作成(ファイルサイズテスト用)."""
        file_path = str(Path(self.temp_dir) / filename)

        # 大量のデータを持つDataFrameを作成
        import numpy as np

        # 約10MB程度のデータを作成
        large_data = {f"column_{i}": np.random.random(50000) for i in range(20)}
        df = pd.DataFrame(large_data)
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)

        return file_path

    def test_file_path_resolution_error_handling(self):
        """ファイルパス解決エラーハンドリングテスト(行217-218)."""
        # 無効なパスでOSErrorを発生させる
        with patch("pathlib.Path.resolve") as mock_resolve:
            mock_resolve.side_effect = OSError("Path resolution failed")

            result = self.loader.is_safe_path("invalid_path")
            assert result is False

        # ValueError を発生させる
        with patch("pathlib.Path.resolve") as mock_resolve:
            mock_resolve.side_effect = ValueError("Invalid path format")

            result = self.loader.is_safe_path("invalid_path")
            assert result is False

    def test_file_size_limit_validation(self):
        """ファイルサイズ制限バリデーションテスト(行250)."""
        test_file = self.create_test_excel("test.xlsx")

        # ファイルサイズを超過させるモック
        with patch("pathlib.Path.stat") as mock_stat:
            mock_stat_result = Mock()
            mock_stat_result.st_size = self.loader.MAX_FILE_SIZE + 1
            mock_stat.return_value = mock_stat_result

            with pytest.raises(ValueError, match="File size .* exceeds limit"):
                self.loader.validate_excel_file(test_file)

    def test_empty_sheet_names_handling(self):
        """空のシート名リスト処理テスト(行275付近)."""
        test_file = self.create_test_excel("test.xlsx")

        with patch("pandas.ExcelFile") as mock_excel_file:
            mock_instance = Mock()
            mock_instance.sheet_names = []  # 空のシート名リスト
            mock_excel_file.return_value = mock_instance

            with pytest.raises(ValueError, match="No sheets found in Excel file"):
                self.loader.basic_sheet_detection(test_file)

    def test_excel_file_reading_errors(self):
        """Excelファイル読み込みエラーテスト."""
        test_file = self.create_test_excel("test.xlsx")

        # pd.read_excelでエラーを発生させる
        with patch("pandas.read_excel") as mock_read_excel:
            mock_read_excel.side_effect = ValueError("Failed to read Excel file")

            with pytest.raises(ValueError):
                self.loader.load_from_excel(test_file)

    def test_invalid_range_specification_formats(self):
        """無効な範囲指定形式のテスト."""
        # Noneや数値の範囲指定
        with pytest.raises(TypeError):
            self.loader._parse_range_specification(None)

        # 空文字列の範囲指定
        with pytest.raises(ValueError):
            self.loader._parse_range_specification("")

        # 無効な形式
        with pytest.raises((ValueError, RangeSpecificationError)):
            self.loader._parse_range_specification("INVALID_FORMAT")

    def test_merge_cells_processing_edge_cases(self):
        """結合セル処理のエッジケーステスト."""
        test_file = self.create_test_excel("merged.xlsx")

        # 結合セル情報が空の場合
        with patch.object(self.loader, "detect_merged_cells") as mock_detect:
            mock_detect.return_value = {
                "has_merged_cells": False,
                "merged_ranges": [],
                "merged_count": 0,
            }

            result = self.loader.load_from_excel_with_merge_cells(
                test_file, merge_mode="expand"
            )
            assert "data" in result
            assert "merged_cells_info" in result

    def test_header_detection_statistical_edge_cases(self):
        """統計的ヘッダー検出のエッジケーステスト."""
        # 全て同じ型のデータ(ヘッダーではない)
        is_header = self.loader._is_likely_header_statistical([1, 2, 3, 4, 5])
        assert is_header is False

        # 混合型だが数値が多い(ヘッダーではない可能性)
        is_header = self.loader._is_likely_header_statistical([1, 2, "text", 4, 5])
        assert is_header is False

        # 空のリスト
        is_header = self.loader._is_likely_header_statistical([])
        assert is_header is False

    def test_column_letter_conversion_extreme_values(self):
        """列文字変換の極値テスト."""
        # 非常に大きな列番号
        large_col = 1000
        letter = self.loader._number_to_column_letter(large_col)
        assert isinstance(letter, str)
        assert len(letter) > 0

        # 変換の一貫性チェック
        converted_back = self.loader._column_letter_to_number(letter)
        assert converted_back == large_col

    def test_data_type_conversion_edge_cases(self):
        """データ型変換のエッジケーステスト."""
        test_file = self.create_test_excel("types.xlsx")

        # 特殊なデータ型を含むExcelファイルをモック
        with patch("pandas.read_excel") as mock_read_excel:
            import numpy as np

            mock_df = pd.DataFrame(
                {
                    "mixed_col": [1, "text", None, np.nan, True, 0.5],
                    "date_col": pd.date_range("2023-01-01", periods=6),
                    "numeric_col": [1, 2, 3, 4, 5, 6],
                }
            )
            mock_read_excel.return_value = mock_df

            result = self.loader.load_from_excel(test_file)
            assert "data" in result
            assert len(result["data"]) > 0

    def test_cache_functionality_edge_cases(self):
        """キャッシュ機能のエッジケーステスト."""
        test_file = self.create_test_excel("cache_test.xlsx")

        # キャッシュディレクトリが存在しない場合
        # キャッシュディレクトリを作成せずにキャッシュ機能をテスト
        result = self.loader.load_from_excel_with_cache(test_file)
        assert "data" in result

        # キャッシュクリア(存在しないファイル)
        nonexistent_file = str(Path(self.temp_dir) / "nonexistent.xlsx")
        self.loader.clear_cache(nonexistent_file)  # エラーが発生しないことを確認

    def test_performance_monitoring_edge_cases(self):
        """パフォーマンス監視のエッジケーステスト."""
        test_file = self.create_test_excel("perf_test.xlsx")

        # メモリ使用量が取得できない場合のモック
        with patch("psutil.Process") as mock_process:
            mock_process.side_effect = ImportError("psutil not available")

            # psutilが利用できない環境でも動作することを確認
            result = self.loader.load_from_excel(test_file)
            assert "data" in result

    def test_range_validation_boundary_conditions(self):
        """範囲検証の境界条件テスト."""
        # 開始行が終了行より大きい
        with pytest.raises((ValueError, RangeSpecificationError)):
            self.loader._validate_range_bounds(
                start_row=10, start_col=1, end_row=5, end_col=5
            )

        # 開始列が終了列より大きい
        with pytest.raises((ValueError, RangeSpecificationError)):
            self.loader._validate_range_bounds(
                start_row=1, start_col=10, end_row=5, end_col=5
            )

        # 負の値
        with pytest.raises((ValueError, RangeSpecificationError)):
            self.loader._validate_range_bounds(
                start_row=-1, start_col=1, end_row=5, end_col=5
            )

    def test_skip_rows_edge_cases(self):
        """行スキップ機能のエッジケーステスト."""
        test_file = self.create_test_excel("skip_test.xlsx")

        # 無効なskip_rows指定
        with pytest.raises((ValueError, TypeError)):
            self.loader.load_from_excel_with_skip_rows(test_file, skip_rows="invalid")

        # データ行数より多いスキップ行数(エラーが発生する)
        with pytest.raises(ValueError, match="Skip row .* is out of range"):
            self.loader.load_from_excel_with_skip_rows(
                test_file,
                skip_rows="10",  # データより多い行数
            )

    def test_multiple_header_rows_edge_cases(self):
        """複数ヘッダー行のエッジケーステスト."""
        test_file = self.create_test_excel("multi_header.xlsx")

        # データ行数より多いヘッダー行数
        with pytest.raises((ValueError, IndexError)):
            self.loader.load_from_excel_with_multiple_headers(
                test_file,
                header_rows=10,  # データより多い行数
            )

        # 負のヘッダー行数
        with pytest.raises(ValueError):
            self.loader.load_from_excel_with_multiple_headers(test_file, header_rows=-1)

    def test_auto_range_detection_edge_cases(self):
        """自動範囲検出のエッジケーステスト."""
        test_file = self.create_test_excel("auto_range.xlsx")

        # 無効な検出モード
        with pytest.raises(ValueError):
            self.loader.load_from_excel_with_detect_range(
                test_file, detect_mode="invalid_mode"
            )

        # 空のExcelファイル(エラーが発生する)
        with patch("pandas.read_excel") as mock_read_excel:
            mock_read_excel.return_value = pd.DataFrame()  # 空のDataFrame

            with pytest.raises(ValueError, match="Empty data in sheet"):
                self.loader.load_from_excel_with_detect_range(
                    test_file, detect_mode="auto"
                )
