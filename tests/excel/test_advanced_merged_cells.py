"""Advanced Merged Cells Testing: 複雑なセル結合シナリオの包括テスト.

ユーザーが実際に使用する複雑な結合パターンを想定した品質保証テスト
- 重なり結合・極端サイズ・データ型混合・エンコーディング問題
- 実世界の複雑な構造(階層・不規則パターン)
- エラー回復・頑健性・メモリ制約
"""

import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Excel対応がある場合のみテストを実行
try:
    from sphinxcontrib.jsontable.excel_data_loader import ExcelDataLoader

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class TestAdvancedMergedCells:
    """Advanced Merged Cells Processing: 高度なセル結合テスト."""

    def setup_method(self):
        """各テストメソッドの前に実行."""
        self.temp_dir = tempfile.mkdtemp()
        self.loader = ExcelDataLoader(self.temp_dir)

    def teardown_method(self):
        """各テストメソッドの後に実行."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def create_overlapping_merges_excel(self) -> str:
        """重なり・交差する結合セルのExcelファイル作成.

        品質リスク: 結合範囲の衝突・重複での予期しない動作
        """
        filename = "overlapping_merges.xlsx"
        file_path = Path(self.temp_dir) / filename
        wb = Workbook()
        ws = wb.active
        ws.title = "OverlappingMerges"

        # シナリオ1: 隣接する結合セル(境界を共有)
        ws["A1"] = "結合1"
        ws.merge_cells("A1:B2")
        ws["C1"] = "結合2"
        ws.merge_cells("C1:D2")  # B列とC列で隣接

        # シナリオ2: L字型の結合パターン
        ws["A4"] = "L字結合1"
        ws.merge_cells("A4:C4")  # 水平結合
        ws["A5"] = "L字結合2"
        ws.merge_cells("A5:A7")  # 垂直結合(A4と結合点共有)

        # シナリオ3: 入れ子状の結合(大結合内に小結合)
        ws["F1"] = "大結合範囲"
        ws.merge_cells("F1:J5")  # 5x5の大結合
        # 注意: 入れ子結合は通常Excel上で無効だが、破損ファイルで発生可能

        # シナリオ4: 十字交差パターン
        ws["A10"] = "水平結合"
        ws.merge_cells("A10:E10")  # 水平結合
        ws["C8"] = "垂直結合"
        ws.merge_cells("C8:C12")  # 垂直結合(C10で交差)

        wb.save(file_path)
        return file_path

    def create_extreme_size_merges_excel(self) -> str:
        """極端なサイズの結合セルExcelファイル作成.

        品質リスク: メモリ使用量・処理時間の異常増大
        """
        filename = "extreme_size_merges.xlsx"
        file_path = Path(self.temp_dir) / filename
        wb = Workbook()
        ws = wb.active

        # シナリオ1: 巨大結合セル(20x20)
        ws["A1"] = "巨大結合データ: " + "X" * 1000  # 大量テキスト
        end_col = get_column_letter(20)
        ws.merge_cells(f"A1:{end_col}20")

        # シナリオ2: 極細長結合(1x100)
        ws["A22"] = "極細長結合"
        ws.merge_cells("A22:A122")

        # シナリオ3: 極幅広結合(100x1)
        ws["A124"] = "極幅広結合"
        end_col = get_column_letter(100)
        ws.merge_cells(f"A124:{end_col}124")

        # シナリオ4: 多数の小結合(メモリテスト)
        for i in range(50):
            cell = f"A{150 + i * 2}"
            ws[cell] = f"小結合{i}"
            next_cell = f"B{150 + i * 2}"
            ws.merge_cells(f"{cell}:{next_cell}")

        wb.save(file_path)
        return file_path

    def create_mixed_datatype_merges_excel(self) -> str:
        """データ型混合の結合セルExcelファイル作成.

        品質リスク: 型変換・値の欠損・フォーマット破損
        """
        filename = "mixed_datatype_merges.xlsx"
        file_path = Path(self.temp_dir) / filename
        wb = Workbook()
        ws = wb.active

        # シナリオ1: 数値と文字列の混合
        ws["A1"] = 123.45  # 数値
        ws["B1"] = "文字列データ"
        ws.merge_cells("A1:B1")

        # シナリオ2: 日付と文字列の混合
        ws["A3"] = "2025-06-13"  # 日付文字列
        ws["B3"] = "テキスト"
        ws.merge_cells("A3:B3")

        # シナリオ3: NULL値との混合
        ws["A5"] = "有効データ"
        ws["B5"] = None  # NULL値
        ws["C5"] = ""  # 空文字
        ws.merge_cells("A5:C5")

        # シナリオ4: 数式を含む結合
        ws["A7"] = "=SUM(1,2,3)"  # 数式
        ws["B7"] = "通常テキスト"
        ws.merge_cells("A7:B7")

        # シナリオ5: 特殊文字・Unicode文字
        ws["A9"] = "🎯💯📊"  # 絵文字
        ws["B9"] = "改行\nテスト"  # 改行文字
        ws["C9"] = "\t\r\n"  # 制御文字
        ws.merge_cells("A9:C9")

        wb.save(file_path)
        return file_path

    def create_hierarchical_structure_excel(self) -> str:
        """階層構造の複雑な結合パターンExcelファイル作成.

        実世界想定: 組織図・会計レポート・統計表の複雑構造
        """
        filename = "hierarchical_structure.xlsx"
        file_path = Path(self.temp_dir) / filename
        wb = Workbook()
        ws = wb.active

        # 会計レポート風の階層構造
        # レベル1: 大項目
        ws["A1"] = "売上高"
        ws.merge_cells("A1:E1")

        # レベル2: 中項目
        ws["A2"] = "商品売上"
        ws.merge_cells("A2:C2")
        ws["D2"] = "サービス売上"
        ws.merge_cells("D2:E2")

        # レベル3: 小項目
        ws["A3"] = "商品A"
        ws["B3"] = "商品B"
        ws["C3"] = "商品C"
        ws["D3"] = "サービス1"
        ws["E3"] = "サービス2"

        # レベル4: データ行
        ws["A4"] = "1000000"
        ws["B4"] = "800000"
        ws["C4"] = "600000"
        ws["D4"] = "500000"
        ws["E4"] = "300000"

        # 複雑な合計セクション
        ws["A6"] = "合計"
        ws.merge_cells("A6:B6")
        ws["C6"] = "小計"
        ws["D6"] = "2400000"
        ws["E6"] = "800000"

        # 総合計行
        ws["A7"] = "総合計"
        ws.merge_cells("A7:D7")
        ws["E7"] = "3200000"

        wb.save(file_path)
        return file_path

    def create_irregular_pattern_excel(self) -> str:
        """不規則パターンの結合セルExcelファイル作成.

        実世界想定: ピボットテーブル風・統計レポート風の不規則構造
        """
        filename = "irregular_pattern.xlsx"
        file_path = Path(self.temp_dir) / filename
        wb = Workbook()
        ws = wb.active

        # ピボットテーブル風の不規則結合
        # 地域別・商品別の入れ子構造
        ws["A1"] = "地域"
        ws.merge_cells("A1:A2")
        ws["B1"] = "商品カテゴリ"
        ws.merge_cells("B1:D1")
        ws["E1"] = "合計"
        ws.merge_cells("E1:E2")

        # 商品カテゴリのサブヘッダー
        ws["B2"] = "商品A"
        ws["C2"] = "商品B"
        ws["D2"] = "商品C"

        # 東京地域データ
        ws["A3"] = "東京"
        ws.merge_cells("A3:A5")  # 3行にわたる結合
        ws["B3"] = "100"
        ws["C3"] = "200"
        ws["D3"] = "150"
        ws["E3"] = "450"

        # 大阪地域データ(不規則な結合)
        ws["A6"] = "大阪"
        ws.merge_cells("A6:A7")  # 2行結合
        ws["B6"] = "80"
        ws["C6"] = "160"
        ws["D6"] = "120"
        ws["E6"] = "360"

        # 名古屋地域(1行のみ)
        ws["A8"] = "名古屋"
        ws["B8"] = "60"
        ws["C8"] = "120"
        ws["D8"] = "90"
        ws["E8"] = "270"

        wb.save(file_path)
        return file_path

    def test_overlapping_merges_expand_mode(self):
        """重なり結合セルのexpandモード処理テスト."""
        excel_path = self.create_overlapping_merges_excel()

        result = self.loader.load_from_excel_with_merge_cells(
            excel_path, merge_mode="expand"
        )

        # 重なり結合が適切に処理されることを確認
        assert result["merge_mode"] == "expand"
        assert result["has_merged_cells"]
        assert len(result["data"]) >= 10  # 十分な行数

        # 隣接結合の展開確認
        assert result["data"][0][0] == "結合1"  # A1
        assert result["data"][0][1] == "結合1"  # B1(展開)
        assert result["data"][0][2] == "結合2"  # C1
        assert result["data"][0][3] == "結合2"  # D1(展開)

    def test_extreme_size_merges_performance(self):
        """極端サイズ結合セルのパフォーマンステスト."""
        excel_path = self.create_extreme_size_merges_excel()

        import time

        start_time = time.time()

        result = self.loader.load_from_excel_with_merge_cells(
            excel_path, merge_mode="expand", header_row=-1
        )

        end_time = time.time()
        processing_time = end_time - start_time

        # パフォーマンス要件(5秒以内)
        assert processing_time < 5.0, f"Processing took {processing_time:.2f} seconds"

        # 巨大結合セルが適切に展開されることを確認
        assert result["merge_mode"] == "expand"
        assert result["has_merged_cells"]

        # 20x20の結合セルが展開されている
        for row in range(min(20, len(result["data"]))):
            for col in range(
                min(20, len(result["data"][row]) if row < len(result["data"]) else 0)
            ):
                if row < len(result["data"]) and col < len(result["data"][row]):
                    cell_value = str(result["data"][row][col])
                    assert "巨大結合データ" in cell_value, (
                        f"Expected '巨大結合データ' in cell at [{row}][{col}], got: '{cell_value[:50]}...'"
                    )

    def test_mixed_datatype_merges_handling(self):
        """データ型混合結合セルの処理テスト."""
        excel_path = self.create_mixed_datatype_merges_excel()

        # 各モードでのデータ型混合処理
        for mode in ["expand", "ignore", "first-value"]:
            result = self.loader.load_from_excel_with_merge_cells(
                excel_path, merge_mode=mode
            )

            assert result["merge_mode"] == mode
            assert result["has_merged_cells"]

            # データ型が文字列に統一されていることを確認
            for row in result["data"]:
                for cell in row:
                    assert isinstance(cell, str) or cell is None

    def test_hierarchical_structure_processing(self):
        """階層構造の複雑結合セル処理テスト."""
        excel_path = self.create_hierarchical_structure_excel()

        result = self.loader.load_from_excel_with_merge_cells(
            excel_path, merge_mode="expand"
        )

        # 階層構造が適切に展開されることを確認
        assert result["data"][0][0] == "売上高"  # A1
        assert result["data"][0][1] == "売上高"  # B1(展開)
        assert result["data"][0][2] == "売上高"  # C1(展開)
        assert result["data"][0][3] == "売上高"  # D1(展開)
        assert result["data"][0][4] == "売上高"  # E1(展開)

        # 中項目レベルの展開確認
        assert result["data"][1][0] == "商品売上"  # A2
        assert result["data"][1][1] == "商品売上"  # B2(展開)
        assert result["data"][1][2] == "商品売上"  # C2(展開)
        assert result["data"][1][3] == "サービス売上"  # D2
        assert result["data"][1][4] == "サービス売上"  # E2(展開)

    def test_irregular_pattern_processing(self):
        """不規則パターン結合セル処理テスト."""
        excel_path = self.create_irregular_pattern_excel()

        result = self.loader.load_from_excel_with_merge_cells(
            excel_path, merge_mode="expand"
        )

        # 不規則な結合パターンが適切に処理されることを確認
        assert result["merge_mode"] == "expand"
        assert result["has_merged_cells"]

        # 東京地域の3行結合確認
        assert result["data"][2][0] == "東京"  # A3
        assert result["data"][3][0] == "東京"  # A4(展開)
        assert result["data"][4][0] == "東京"  # A5(展開)

    def test_memory_constraint_handling(self):
        """メモリ制約下での結合セル処理テスト."""
        excel_path = self.create_extreme_size_merges_excel()

        # メモリ使用量の監視(tracemalloc版)
        import tracemalloc

        tracemalloc.start()

        result = self.loader.load_from_excel_with_merge_cells(
            excel_path, merge_mode="expand"
        )

        current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()

        # ピークメモリ使用量が50MB以内であることを確認
        assert peak < 50 * 1024 * 1024, f"Peak memory usage: {peak / 1024 / 1024:.1f}MB"

        assert result["has_merged_cells"]

    def test_error_recovery_corrupted_merges(self):
        """破損した結合セル定義でのエラー回復テスト."""
        # 正常なファイルを作成してから手動で破損させるのは困難なため、
        # 無効な範囲指定での処理をテスト
        excel_path = self.create_overlapping_merges_excel()

        # 範囲外を指定した場合のエラー回復
        try:
            result = self.loader.load_from_excel_with_merge_cells_and_range(
                excel_path, range_spec="Z100:AA200", merge_mode="expand"
            )
            # エラーが発生せず、空のデータが返されることを確認
            assert result["data"] == []
        except Exception as e:
            # 適切なエラーメッセージが含まれることを確認
            assert "range" in str(e).lower() or "invalid" in str(e).lower()

    def test_concurrent_merge_processing(self):
        """並行処理での結合セル処理テスト."""
        import concurrent.futures

        excel_path = self.create_hierarchical_structure_excel()
        results = []
        errors = []

        def process_merge_cells(mode):
            try:
                result = self.loader.load_from_excel_with_merge_cells(
                    excel_path, merge_mode=mode
                )
                return result
            except Exception as e:
                errors.append(e)
                return None

        # 複数スレッドで同時実行
        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
            futures = [
                executor.submit(process_merge_cells, "expand"),
                executor.submit(process_merge_cells, "ignore"),
                executor.submit(process_merge_cells, "first-value"),
            ]

            for future in concurrent.futures.as_completed(futures):
                result = future.result()
                if result:
                    results.append(result)

        # 並行処理でエラーが発生しないことを確認
        assert len(errors) == 0, f"Concurrent processing errors: {errors}"
        assert len(results) == 3, "All concurrent processes should complete"

        # 各モードで正しい結果が得られることを確認
        modes = [r["merge_mode"] for r in results]
        assert "expand" in modes
        assert "ignore" in modes
        assert "first-value" in modes


if __name__ == "__main__":
    # スタンドアロンテスト実行
    pytest.main([__file__, "-v"])
