#!/usr/bin/env python3
"""ヘッダー行機能のデバッグスクリプト"""

import os
import tempfile
import shutil
from openpyxl import Workbook
from sphinxcontrib.jsontable.excel_data_loader import ExcelDataLoader

def create_header_test_excel() -> str:
    """ヘッダー行設定テスト用のExcelファイルを作成"""
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, "header_test.xlsx")

    # 複数行でヘッダーが異なる位置にあるデータを作成
    data = [
        ["メタデータ", "作成日: 2025-06-13", "", ""],  # Row 0: メタデータ行
        ["説明", "売上データの月次集計", "", ""],  # Row 1: 説明行
        ["", "", "", ""],  # Row 2: 空行
        ["商品名", "1月売上", "2月売上", "3月売上"],  # Row 3: ヘッダー行
        ["商品A", "100000", "120000", "110000"],  # Row 4: データ行
        ["商品B", "150000", "180000", "160000"],  # Row 5: データ行
        ["商品C", "80000", "90000", "85000"],  # Row 6: データ行
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # データを行ごとに書き込み
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(file_path)
    return file_path, temp_dir

def main():
    """メイン関数"""
    # テスト用Excelファイルを作成
    excel_path, temp_dir = create_header_test_excel()
    
    print(f"Created Excel file: {excel_path}")
    
    # 生のExcelデータを確認
    print("\n=== 生のExcelデータ（openpyxlで直読み） ===")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        print("Raw Excel data:")
        for row_idx in range(1, 8):  # 1-7行目
            row_data = []
            for col_idx in range(1, 5):  # A-D列
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            print(f"  Excel Row {row_idx}: {row_data}")
    except Exception as e:
        print(f"Error reading raw Excel: {e}")
    
    # ExcelDataLoaderを初期化
    loader = ExcelDataLoader(temp_dir)
    
    # 通常の読み込み
    print("\n=== 通常読み込み ===")
    try:
        result_normal = loader.load_from_excel(excel_path)
        print(f"Normal result keys: {list(result_normal.keys())}")
        print(f"Normal data shape: {len(result_normal['data'])} rows")
        print("All normal data:")
        for i, row in enumerate(result_normal['data']):
            print(f"  Row {i}: {row}")
    except Exception as e:
        print(f"Error in normal load: {type(e).__name__}: {e}")
    
    # ヘッダー行指定読み込み（テスト値）
    print("\n=== ヘッダー行指定読み込み (header_row=2 - empty row) ===")
    try:
        result_header = loader.load_from_excel_with_header_row(excel_path, header_row=2)
        print(f"Headers: {result_header.get('headers', 'NOT_FOUND')}")
        print(f"Data rows: {len(result_header.get('data', []))}")
    except Exception as e:
        print(f"Error: {e}")

    print("\n=== ヘッダー行指定読み込み (header_row=3 - actual header) ===")
    try:
        result_header = loader.load_from_excel_with_header_row(excel_path, header_row=3)
        print(f"Headers: {result_header.get('headers', 'NOT_FOUND')}")
        print(f"Data rows: {len(result_header.get('data', []))}")
    except Exception as e:
        print(f"Error: {e}")
    
    # クリーンアップ
    shutil.rmtree(temp_dir, ignore_errors=True)

if __name__ == "__main__":
    main()